# Code/main.py

import os
import sys
import pandas as pd


from config import FRED_API_KEY

from loader import load_company_facts, get_cik_from_ticker
from agents.filing_agent import analyze_filing, extract_margin_series
from agents.margin_diagnostic_agent import run_margin_diagnostic
from agents.working_capital_agent import run_working_capital_engine
from agents.risk_scoring_agent import run_risk_scoring_layer
from agents.event_engine_agent import run_event_engine
from agents.event_impact_agent import aggregate_event_impacts
from agents.valuation_agent import run_scenario_dcf, run_sensitivity_analysis
from agents.signal_agent import generate_investment_signal, get_market_data
from agents.memo_agent import generate_investment_memo

try:
    from agents.data_inputs_agent import build_data_inputs
    HAS_DATA_INPUTS = True
except:
    HAS_DATA_INPUTS = False

def human_format(num):
    if num is None:
        return None

    try:
        num = float(num)
    except:
        return num

    if abs(num) >= 1e12:
        return f"{num/1e12:.2f}T"
    elif abs(num) >= 1e9:
        return f"{num/1e9:.2f}B"
    elif abs(num) >= 1e6:
        return f"{num/1e6:.2f}M"
    elif abs(num) >= 1e3:
        return f"{num/1e3:.2f}K"
    else:
        return round(num, 4)

def to_vertical_df(data_dict):

    # Strict: only handle dicts properly
    if isinstance(data_dict, dict):
        return pd.DataFrame(
            list(data_dict.items()),
            columns=["Metric", "Value"]
        )

    # If something went wrong, show FULL debug
    return pd.DataFrame(
        [{"Metric": "DEBUG", "Value": repr(data_dict)}]
    )

def run_analysis(ticker):

    os.makedirs("outputs", exist_ok=True)

    print(f"\n===== RUNNING WORKFLOW FOR {ticker.upper()} =====")

    # ----------------------------------
    # 1️⃣ Pull EDGAR Structured Data
    # ----------------------------------
    facts_json = load_company_facts(ticker)

    # ----------------------------------
    # 2️⃣ Snapshot Financial Metrics
    # ----------------------------------
    snapshot_metrics = analyze_filing(facts_json)
    print("Snapshot metrics computed.")

    # ----------------------------------
    # 3️⃣ Margin Diagnostic
    # ----------------------------------
    margin_series = extract_margin_series(facts_json)
    margin_report = run_margin_diagnostic(margin_series, "")

    # ----------------------------------
    # 🔗 Inject REAL Operating Metrics
    # ----------------------------------
    if margin_series and len(margin_series) > 0:

        latest_margin = margin_series[-1]["operating_margin"]

        snapshot_metrics["Operating Margin"] = latest_margin

        revenue = snapshot_metrics.get("Revenue")

        if revenue and latest_margin:
            reconstructed_income = revenue * latest_margin
            snapshot_metrics["Operating Income"] = reconstructed_income

        print("Injected Operating Margin:", latest_margin)
        print("Reconstructed Operating Income:", reconstructed_income)

    else:
        print("⚠️ No margin series available")
        
    # ----------------------------------
    # 4️⃣ Working Capital
    # ----------------------------------
    wc_report = run_working_capital_engine(
        facts_json,
        snapshot_metrics
    )
    print("Working capital diagnostic computed.")

    # ----------------------------------
    # 5️⃣ Event Intelligence
    # ----------------------------------
    cik = get_cik_from_ticker(ticker)
    event_report = run_event_engine(cik)
    print("Event intelligence computed.")

    # ----------------------------------
    # 6️⃣ Event Impact
    # ----------------------------------
    event_impact = aggregate_event_impacts(event_report)

    print("Event impact computed.")

    # ----------------------------------
    # 7️⃣ Risk Scoring
    # ----------------------------------
    risk_report = run_risk_scoring_layer(
        margin_report,
        wc_report,
        snapshot_metrics
    )
    print("Risk scoring computed.")

    # ----------------------------------
    # 8️⃣ Data Inputs
    # ----------------------------------
    data_inputs = None

    if HAS_DATA_INPUTS:
        try:
            data_inputs = build_data_inputs(
                ticker,
                snapshot_metrics,
                margin_report,
                risk_report,
                FRED_API_KEY
            )

            print("Data inputs computed.")

        except Exception as e:
            print("❌ DATA INPUTS FAILED:", e)
            raise e
        

    # ----------------------------------
    # 9️⃣ Excel Output
    # ----------------------------------
    output_path = f"outputs/{ticker.upper()}_analysis.xlsx"

    # ----------------------------------
    # 🔟 Valuation Engine
    # ----------------------------------
    scenario_output = {}
    valuation_output = None
    sensitivity_df = None

    market_data = None
    shares_outstanding = None
    if data_inputs:
        market_data = get_market_data(ticker)
        shares_outstanding = market_data.get("Shares Outstanding")

        scenario_output = run_scenario_dcf(
            data_inputs,
            event_impact,
            snapshot_metrics,
            shares_outstanding=shares_outstanding
        )
        valuation_output = scenario_output.get("Base")
        
        sensitivity_df = run_sensitivity_analysis(
            data_inputs,
            event_impact,
            snapshot_metrics,
            shares_outstanding=shares_outstanding
        )

        print("Valuation computed.")

    # ----------------------------------
    # 1️⃣1️⃣ Signal Generation 
    # ----------------------------------
    signal_output = None

    if valuation_output and data_inputs and market_data:

        signal_output = generate_investment_signal(
            valuation_output,
            market_data,
            data_inputs
        )

        print("Signal generated.")

    # ----------------------------------
    # 📝 Investment Memo
    # ----------------------------------
    memo_dict, memo_text = None, None
    if signal_output and scenario_output:
        memo_dict, memo_text = generate_investment_memo(
            signal_output,
            scenario_output,
            risk_report,
            event_report
        )
        if memo_dict:
            print("Investment memo computed.")

    # ----------------------------------
    # 🔗 Combine Valuation + Signal
    # ----------------------------------
    combined_output = None

    if valuation_output and signal_output:

        combined_output = {}

        # --- Signal ---
        combined_output["Signal"] = signal_output.get("Signal")
        combined_output["Confidence Level"] = signal_output.get("Confidence Level")
        combined_output["Upside (%)"] = signal_output.get("Upside (%)")
        combined_output["Market Price"] = signal_output.get("Market Price")
        combined_output["Shares Outstanding"] = market_data.get("Shares Outstanding")
        
        # --- Value Output ---
        combined_output["Enterprise Value"] = valuation_output.get("Enterprise Value")
        combined_output["Net Debt"] = valuation_output.get("Net Debt")
        combined_output["Equity Value"] = valuation_output.get("Equity Value")
        
        combined_output["Intrinsic Value (Total)"] = valuation_output.get("Intrinsic Value")
        combined_output["Intrinsic Value (Per Share)"] = valuation_output.get("Intrinsic Value Per Share")

        # --- Valuation ---
        combined_output["5Y PV Cash Flows"] = valuation_output.get("5Y PV Cash Flows")
        combined_output["Terminal Value"] = valuation_output.get("Terminal Value")
        combined_output["Terminal Value Contribution (%)"] = valuation_output.get("Terminal Value Contribution (%)")

        # --- Assumptions ---
        combined_output["Stage 1 Growth"] = valuation_output.get("Stage 1 Growth")
        combined_output["Discount Used"] = valuation_output.get("Discount Used")
        combined_output["Raw ERP"] = signal_output.get("Raw ERP")
        combined_output["Adjusted ERP"] = signal_output.get("Adjusted ERP")

        # --- Drivers ---
        drivers = signal_output.get("Key Drivers", [])
        for i, d in enumerate(drivers):
            combined_output[f"Driver {i+1}"] = d


    def format_section(ws, start_row, num_rows):

        headers = [cell.value for cell in ws[start_row - 1]]

        for r in ws.iter_rows(min_row=start_row, max_row=start_row + num_rows - 1):

            for col_idx, cell in enumerate(r):

                header = str(headers[col_idx]).lower() if headers[col_idx] else ""

                if isinstance(cell.value, (int, float)):

                    # -----------------------------
                    # YEAR COLUMN (force integer)
                    # -----------------------------
                    if "year" in header:
                        cell.number_format = '0'
                        cell.value = int(cell.value)

                    # -----------------------------
                    # PERCENT FIELDS
                    # -----------------------------
                    elif any(x in header for x in [
                        "growth", "discount", "spread", "rate", "contribution"
                    ]):
                        cell.number_format = '0.00%'

                    # -----------------------------
                    # LARGE NUMBERS
                    # -----------------------------
                    elif abs(cell.value) >= 1_000_000:
                        cell.number_format = '#,##0'

                    # -----------------------------
                    # DEFAULT
                    # -----------------------------
                    else:
                        cell.number_format = '0.00'
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        # ----------------------------------
        # 01 Financial Analysis 
        # ----------------------------------
        sheet_name = "01_Financial_Analysis"
        row = 0

        snapshot_df = to_vertical_df(snapshot_metrics)
        snapshot_df.to_excel(writer, sheet_name=sheet_name, startrow=row, index=False)

        row += len(snapshot_df) + 2

        pd.DataFrame([["--- MARGINS ---"]]).to_excel(
            writer, sheet_name=sheet_name, startrow=row, index=False, header=False
        )
        row += 1

        margin_df = to_vertical_df(margin_report)
        margin_df.to_excel(writer, sheet_name=sheet_name, startrow=row, index=False)

        row += len(margin_df) + 2

        pd.DataFrame([["--- WORKING CAPITAL ---"]]).to_excel(
            writer, sheet_name=sheet_name, startrow=row, index=False, header=False
        )
        row += 1

        wc_df = to_vertical_df(wc_report)
        wc_df.to_excel(writer, sheet_name=sheet_name, startrow=row, index=False)


        # ----------------------------------
        # 02 Event Intelligence
        # ----------------------------------
        pd.DataFrame(event_report).to_excel(
            writer, sheet_name="02_Events", index=False
        )

        # ----------------------------------
        # 03 Risk
        # ----------------------------------
        to_vertical_df(risk_report).to_excel(
            writer, sheet_name="03_Risk", index=False
        )

        # ----------------------------------
        # 04 Data Inputs
        # ----------------------------------
        if data_inputs:
            to_vertical_df(data_inputs).to_excel(
                writer, sheet_name="04_Data_Inputs", index=False
            )


        # ==================================
        # 05 COMBINED VALUATION MODEL
        # ==================================
        if combined_output:

            sheet_name = "05_Valuation_Model"
            row = 0

            from openpyxl.styles import Font, PatternFill
            bold_font = Font(bold=True)

            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

            # ----------------------------------
            # 🔷 1. SIGNAL + VALUATION
            # ----------------------------------
            pd.DataFrame([["--- SIGNAL & VALUATION ---"]]).to_excel(
                writer, sheet_name=sheet_name, startrow=row, index=False, header=False
            )
            row += 1

            summary_df = to_vertical_df(combined_output)
            summary_df.to_excel(writer, sheet_name=sheet_name, startrow=row, index=False)

            ws = writer.sheets[sheet_name]
            ws.cell(row=row, column=1).font = bold_font

            # Format + color
            for r in ws.iter_rows(min_row=row+1, max_row=row+len(summary_df)):
                metric_cell = r[0]
                value_cell = r[1]

                metric = str(metric_cell.value)

                if isinstance(value_cell.value, (int, float)):
                    if any(x in metric for x in ["%", "Growth", "Discount", "Contribution"]):
                        value_cell.number_format = '0.00%'
                    elif abs(value_cell.value) >= 1_000_000:
                        value_cell.number_format = '#,##0'
                    else:
                        value_cell.number_format = '0.00'

                if metric == "Signal":
                    if value_cell.value == "Bullish":
                        value_cell.fill = green_fill
                    elif value_cell.value == "Bearish":
                        value_cell.fill = red_fill
                    else:
                        value_cell.fill = yellow_fill

                if metric == "Upside (%)" and isinstance(value_cell.value, (int, float)):
                    if value_cell.value > 0.10:
                        value_cell.fill = green_fill
                    elif value_cell.value < -0.10:
                        value_cell.fill = red_fill
                    else:
                        value_cell.fill = yellow_fill

            row += len(summary_df) + 2


            # ----------------------------------
            # 🔷 2. EVENT IMPACT
            # ----------------------------------
            pd.DataFrame([["--- EVENT IMPACT ---"]]).to_excel(
                writer, sheet_name=sheet_name, startrow=row, index=False, header=False
            )
            ws.cell(row=row+1, column=1).font = bold_font

            row += 1

            event_df = to_vertical_df(event_impact)
            event_df.to_excel(writer, sheet_name=sheet_name, startrow=row, index=False)
            format_section(ws, row + 1, len(event_df))

            row += len(event_df) + 2


            # ----------------------------------
            # 🔷 3. DIAGNOSTICS
            # ----------------------------------
            pd.DataFrame([["--- DIAGNOSTICS ---"]]).to_excel(
                writer, sheet_name=sheet_name, startrow=row, index=False, header=False
            )
            ws.cell(row=row+1, column=1).font = bold_font

            row += 1

            diag_df = to_vertical_df(valuation_output["Diagnostics"])
            diag_df.to_excel(writer, sheet_name=sheet_name, startrow=row, index=False)
            format_section(ws, row + 1, len(diag_df))

            row += len(diag_df) + 2


            # ----------------------------------
            # 🔷 4. PROJECTION DETAILS
            # ----------------------------------
            pd.DataFrame([["--- PROJECTION DETAILS ---"]]).to_excel(
                writer, sheet_name=sheet_name, startrow=row, index=False, header=False
            )
            ws.cell(row=row+1, column=1).font = bold_font

            row += 1

            # -----------------------------
            # Build DataFrame
            # -----------------------------
            projection_df = pd.DataFrame(valuation_output["Projection Details"])

            # 🔥 FIX 1: enforce correct data types
            projection_df["year"] = projection_df["year"].astype(int)
            projection_df["growth"] = projection_df["growth"].astype(float)
            projection_df["projected_cf"] = projection_df["projected_cf"].astype(float)
            projection_df["discounted_cf"] = projection_df["discounted_cf"].astype(float)

            # -----------------------------
            # Write to Excel
            # -----------------------------
            projection_df.to_excel(
                writer,
                sheet_name=sheet_name,
                startrow=row,
                index=False
            )

            # -----------------------------
            # 🔥 FIX 2: format ONLY data rows (skip header)
            # -----------------------------
            format_section(ws, row + 1, len(projection_df))

            # ==================================
            # 07 SCENARIO ANALYSIS
            # ==================================
            if scenario_output:
                sheet_name_sc = "07_Scenario_Analysis"
                
                # Check if we have market data
                market_price = combined_output.get("Market Price") if combined_output else None
                shares_out = combined_output.get("Shares Outstanding") if combined_output else None

                scenario_rows = []
                for sc_name, sc_val in scenario_output.items():
                    intr_val = sc_val.get("Intrinsic Value")
                    upside = None
                    if market_price and shares_out and intr_val:
                        intr_per_share = intr_val / shares_out
                        upside = (intr_per_share / market_price) - 1

                    scenario_rows.append({
                        "Scenario": sc_name,
                        "Intrinsic Value": intr_val,
                        "Upside": upside,
                        "Growth": sc_val.get("Stage 1 Growth"),
                        "Discount": sc_val.get("Discount Used")
                    })

                sc_df = pd.DataFrame(scenario_rows)
                sc_df.to_excel(writer, sheet_name=sheet_name_sc, index=False)
                
                sc_ws = writer.sheets[sheet_name_sc]
                # Format section takes starting data row (2)
                format_section(sc_ws, 2, len(sc_df))

            # ==================================
            # 08 SENSITIVITY ANALYSIS
            # ==================================
            if sensitivity_df is not None:
                sheet_name_sens = "08_Sensitivity_Analysis"
                
                # Write to Excel
                sensitivity_df.to_excel(writer, sheet_name=sheet_name_sens, index=False)
                
                sens_ws = writer.sheets[sheet_name_sens]
                
                # Ensure headers are bold
                for cell in sens_ws[1]:
                    cell.font = bold_font
                    
                # Format matrix
                for r in sens_ws.iter_rows(min_row=2, max_row=1+len(sensitivity_df)):
                    # First column is Growth Rate -> Format as %
                    r[0].number_format = '0.00%'
                    # Values -> Currency/Decimals
                    for col_idx in range(1, len(r)):
                        val = r[col_idx].value
                        if isinstance(val, (int, float)):
                            if abs(val) >= 1_000_000:
                                r[col_idx].number_format = '$#,##0'
                            else:
                                r[col_idx].number_format = '$#,##0.00'

            # ==================================
            # 09 INVESTMENT MEMO
            # ==================================
            if memo_dict and memo_text:
                sheet_name_memo = "09_Investment_Memo"
                row_idx = 0

                import openpyxl
                from openpyxl.styles import Alignment

                # Top Natural Language Summary
                pd.DataFrame([[memo_text]]).to_excel(
                    writer, sheet_name=sheet_name_memo, startrow=row_idx, index=False, header=False
                )
                
                memo_ws = writer.sheets[sheet_name_memo]
                memo_ws.cell(row=row_idx+1, column=1).font = bold_font
                memo_ws.cell(row=row_idx+1, column=1).alignment = Alignment(wrap_text=True)
                memo_ws.column_dimensions['A'].width = 80
                
                row_idx += 2

                # Dump memo_dict sections
                for section_name, metrics in memo_dict.items():
                    # Section Header
                    pd.DataFrame([[f"--- {section_name.upper()} ---"]]).to_excel(
                        writer, sheet_name=sheet_name_memo, startrow=row_idx, index=False, header=False
                    )
                    memo_ws.cell(row=row_idx+1, column=1).font = bold_font
                    row_idx += 1

                    # Section Data
                    section_df = to_vertical_df(metrics)
                    section_df.to_excel(writer, sheet_name=sheet_name_memo, startrow=row_idx, index=False)
                    format_section(memo_ws, row_idx+1, len(section_df))

                    row_idx += len(section_df) + 2

            print(f"\nExcel report saved: {output_path}")
            print("\n===== ANALYSIS COMPLETE =====")


if __name__ == "__main__":

    if len(sys.argv) < 2:
        print("Usage: python Code/main.py TICKER")
        sys.exit()

    ticker_input = sys.argv[1]
    run_analysis(ticker_input)